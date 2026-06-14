#!/usr/bin/env python3
"""
Marriott vs. Hilton — Washington, DC Hotel Saturation Map

A one-off analytical map (not an STR comp set, not a brand portfolio). Reads a
CoStar export of DC hotels, classifies each property into the Marriott or Hilton
brand family (citizenM falls out as "Other"), derives operating/pipeline/inactive
status from CoStar's Building Status, and renders a standalone Google Maps page so
you can see how densely each operator covers the market relative to a starred
subject — the AKA White House (an independent serviced residence).

Source : STR Drops/Marriott vs. Hilton DC Saturation/CostarExport.xlsx
Output : <repo>/marriott-vs-hilton-dc-saturation/index.html

Usage:
    python generate_dc_saturation_map.py            # generate only
    python generate_dc_saturation_map.py --deploy   # generate + push + verify
"""

import argparse
import json
import math
import os
import sys

import openpyxl

from generate_str_map import (
    REPO_DIR,
    API_KEY,
    PAGES_BASE_URL,
    _commit_and_push,
    verify_deploy,
)

SRC = r"C:\Users\acarr\OneDrive\Documents\STR Drops\Marriott vs. Hilton DC Saturation\CostarExport.xlsx"
SLUG = "marriott-vs-hilton-dc-saturation"

# AKA White House — the starred subject. Independent luxury serviced residence at
# 1710 H St NW; not in the CoStar brand export. Coords already in geocache.json.
AKA = {
    "name": "AKA White House",
    "address": "1710 H St NW, Washington, DC 20006",
    "lat": 38.9000628,
    "lng": -77.0404004,
    "note": "Subject property — luxury serviced residence, two blocks from the "
            "White House. Independent (not a Marriott or Hilton flag).",
}

# Brand-family classification. Lower-cased substring match against the property
# name. Order doesn't matter — a name matching both families would be flagged.
FAMILIES = {
    "Marriott": {
        "color": "#c8102e", "stroke": "#ffffff", "label": "#ffffff",
        "keys": [
            "marriott", "ritz-carlton", "ritz carlton", "st. regis", "st regis",
            "westin", "le meridien", "le méridien", "sheraton", "aloft",
            "element ", "four points", "moxy", "ac hotel", "ac hotels",
            "residence inn", "courtyard", "fairfield", "springhill", "autograph",
            "tribute portfolio", "luxury collection", "renaissance", "gaylord",
            "delta hotels", "protea", "city express", "series by marriott",
        ],
    },
    "Hilton": {
        "color": "#0a4d8c", "stroke": "#ffffff", "label": "#ffffff",
        "keys": [
            "hilton", "hampton", "embassy suites", "canopy", "conrad",
            "waldorf astoria", "curio", "tapestry", "motto", "tempo", "homewood",
            "home2", "doubletree", "tru by hilton", "signia", "lxr",
            "spark by hilton",
        ],
    },
}
OTHER = {"color": "#94a3b8", "stroke": "#ffffff", "label": "#1a2744"}

# CoStar Building Status -> saturation bucket.
STATUS_MAP = {
    "Existing": "operating",
    "Under Construction": "pipeline",
    "Final Planning": "pipeline",
    "Deferred": "inactive",
    "Demolished": "inactive",
    "Abandoned": "inactive",
}


def classify_family(name):
    n = name.lower()
    matches = [fam for fam, cfg in FAMILIES.items()
               if any(k in n for k in cfg["keys"])]
    if len(matches) > 1:
        print(f"  WARNING: '{name}' matched multiple families {matches}; using {matches[0]}")
    return matches[0] if matches else "Other"


def haversine_mi(lat1, lng1, lat2, lng2):
    R = 3958.8  # miles
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


def load_hotels():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    hotels = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        lat = ws.cell(row=r, column=39).value
        lng = ws.cell(row=r, column=40).value
        if lat in (None, "") or lng in (None, ""):
            print(f"  WARNING: '{name}' has no coordinates; skipping")
            continue
        raw_status = ws.cell(row=r, column=6).value or "Existing"
        hotels.append({
            "name": name,
            "family": classify_family(name),
            "status": STATUS_MAP.get(raw_status, "inactive"),
            "raw_status": raw_status,
            "star": ws.cell(row=r, column=3).value,
            "rba": ws.cell(row=r, column=7).value,
            "year": ws.cell(row=r, column=27).value,
            "stories": ws.cell(row=r, column=38).value,
            "lat": float(lat),
            "lng": float(lng),
            "dist": round(haversine_mi(AKA["lat"], AKA["lng"], float(lat), float(lng)), 2),
        })
    # Rank by proximity to the subject — the list doubles as a "nearest flags to AKA".
    hotels.sort(key=lambda h: h["dist"])
    for i, h in enumerate(hotels, 1):
        h["n"] = i
    return hotels


def summarize(hotels):
    """Saturation matrix: family x status hotel counts, plus operating RBA."""
    fams = ["Marriott", "Hilton", "Other"]
    states = ["operating", "pipeline", "inactive"]
    matrix = {f: {s: 0 for s in states} for f in fams}
    op_rba = {f: 0 for f in fams}
    for h in hotels:
        matrix[h["family"]][h["status"]] += 1
        if h["status"] == "operating" and isinstance(h["rba"], (int, float)):
            op_rba[h["family"]] += h["rba"]
    return {"matrix": matrix, "op_rba": op_rba, "fams": fams, "states": states}


# ---------------------------------------------------------------------------
# HTML template. CSS/JS kept as a plain string (placeholders via .replace) so we
# don't fight f-string brace escaping across a large script block. Visual
# language borrowed from generate_brand_map.py.
# ---------------------------------------------------------------------------
TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Marriott vs. Hilton &mdash; Washington, DC Hotel Saturation</title>
  <style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { font-family: 'Segoe UI', Arial, sans-serif; height: 100vh; display: flex; flex-direction: column; background: #f5f5f5; }
    .header { background: #1a2744; border-bottom: 3px solid #c8102e; padding: 14px 22px; display: flex; align-items: center; justify-content: space-between; flex-shrink: 0; }
    .header-left h1 { font-size: 18px; font-weight: 700; color: #fff; }
    .header-left p { font-size: 11px; color: #8fa8d8; margin-top: 3px; }
    .header-right { font-size: 11px; color: #8fa8d8; }
    .header-right a { color: #8fa8d8; text-decoration: none; }
    .header-right a:hover { color: #fff; }
    .main { display: flex; flex: 1; overflow: hidden; }
    .sidebar { width: 340px; background: #fff; border-right: 1px solid #dde3ed; overflow-y: auto; flex-shrink: 0; box-shadow: 2px 0 8px rgba(0,0,0,0.06); }
    .sb-section { padding: 14px; border-bottom: 1px solid #eaecf0; }
    .sb-section h3 { font-size: 10px; text-transform: uppercase; letter-spacing: 0.9px; color: #8090b0; margin-bottom: 10px; font-weight: 600; }
    .legend-row { display: flex; align-items: center; gap: 8px; margin-bottom: 2px; font-size: 12.5px; color: #334; padding: 5px 6px; margin-left: -6px; margin-right: -6px; border-radius: 4px; }
    .legend-row.clickable { cursor: pointer; user-select: none; }
    .legend-row.clickable:hover { background: #f0f5ff; }
    .legend-row.off { opacity: 0.35; }
    .legend-row.off .leg-dot { background: #cbd5e1 !important; border-color: #cbd5e1 !important; }
    .legend-row .leg-count { margin-left: auto; color: #8090b0; font-variant-numeric: tabular-nums; font-weight: 600; }
    .leg-dot { width: 14px; height: 14px; transform: rotate(45deg); border: 2px solid rgba(0,0,0,0.15); flex-shrink: 0; }
    .leg-star { width: 18px; height: 18px; flex-shrink: 0; color: #f4c20d; filter: drop-shadow(0 1px 1px rgba(0,0,0,0.3)); }
    .filter-pills { display: flex; flex-wrap: wrap; gap: 6px; }
    .filter-pill { display: inline-flex; align-items: center; gap: 6px; padding: 3px 9px; border-radius: 12px; font-size: 10.5px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.4px; cursor: pointer; user-select: none; border: 1.5px solid; background: #fff; }
    .filter-pill.off { opacity: 0.35; background: #f3f4f6; }
    .filter-pill .pill-count { font-weight: 400; opacity: 0.7; font-variant-numeric: tabular-nums; letter-spacing: 0; }
    .filter-pill.s-operating { color: #047857; border-color: #047857; }
    .filter-pill.s-pipeline  { color: #b45309; border-color: #b45309; }
    .filter-pill.s-inactive  { color: #6b7280; border-color: #6b7280; }
    .fit-btn { width: 100%; margin-top: 10px; padding: 6px 10px; font-size: 11px; font-weight: 600; color: #1a2744; background: #f0f5ff; border: 1px solid #c8d4e8; border-radius: 4px; cursor: pointer; }
    .fit-btn:hover { background: #e0eaff; }
    .filter-note { font-size: 10px; color: #8090b0; margin-top: 8px; font-style: italic; line-height: 1.4; }
    .matrix { width: 100%; border-collapse: collapse; font-size: 12px; font-variant-numeric: tabular-nums; }
    .matrix th, .matrix td { padding: 5px 6px; text-align: right; }
    .matrix th:first-child, .matrix td:first-child { text-align: left; }
    .matrix thead th { font-size: 9.5px; color: #fff; text-transform: uppercase; letter-spacing: 0.4px; font-weight: 700; }
    .matrix thead th.c-fam { background: #1a2744; border-top-left-radius: 4px; }
    .matrix thead th.s-operating { background: #047857; }
    .matrix thead th.s-pipeline  { background: #b45309; }
    .matrix thead th.s-inactive  { background: #6b7280; }
    .matrix thead th.s-total     { background: #1a2744; border-top-right-radius: 4px; }
    .matrix tbody td { border-bottom: 1px solid #eef0f3; }
    .matrix tbody tr td:first-child { font-weight: 600; color: #1a2744; }
    .matrix .fam-dot { display: inline-block; width: 9px; height: 9px; transform: rotate(45deg); margin-right: 6px; vertical-align: middle; border: 1px solid rgba(0,0,0,0.15); }
    .matrix td.tot { font-weight: 700; color: #1a2744; }
    .matrix td.dash { color: #d1d5db; }
    .matrix tfoot td { font-weight: 800; color: #1a2744; border-top: 2px solid #1a2744; padding-top: 7px; }
    .matrix-note { font-size: 10px; color: #8090b0; margin-top: 9px; line-height: 1.45; font-style: italic; }
    .rba-row { display: flex; justify-content: space-between; font-size: 11.5px; padding: 3px 0; color: #334; }
    .rba-row .rba-lab { display: flex; align-items: center; gap: 7px; }
    .hotel-item { display: flex; align-items: flex-start; gap: 9px; padding: 9px 6px; cursor: pointer; border-bottom: 1px solid #f0f2f5; }
    .hotel-item:hover { background: #f0f5ff; }
    .hotel-item.off { display: none; }
    .h-pin { width: 22px; height: 22px; transform: rotate(45deg); display: flex; align-items: center; justify-content: center; flex-shrink: 0; margin-top: 2px; }
    .h-pin-num { transform: rotate(-45deg); color: #fff; font-weight: 700; font-size: 10px; }
    .h-info h4 { font-size: 12px; font-weight: 600; color: #1a2744; line-height: 1.3; }
    .h-info .h-meta { font-size: 10px; color: #7080a0; margin-top: 2px; }
    .h-status-badge { font-size: 8px; padding: 1px 4px; border-radius: 2px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.3px; margin-left: 4px; vertical-align: middle; }
    .h-status-badge.operating { background: #d1fae5; color: #047857; }
    .h-status-badge.pipeline  { background: #fef3c7; color: #b45309; }
    .h-status-badge.inactive  { background: #f3f4f6; color: #6b7280; }
    .sidebar-footer { padding: 10px 14px; font-size: 10px; color: #aabbd0; border-top: 1px solid #eaecf0; line-height: 1.5; }
    #map { flex: 1; }
    .gm-iw { font-family: 'Segoe UI', Arial, sans-serif; min-width: 230px; max-width: 290px; }
    .gm-iw-title { font-size: 14px; font-weight: 700; color: #1a2744; line-height: 1.3; }
    .gm-iw-brand { font-size: 10px; color: #fff; display: inline-block; padding: 2px 7px; border-radius: 3px; text-transform: uppercase; letter-spacing: 0.6px; margin: 5px 6px 8px 0; font-weight: 700; }
    .gm-iw-status { display: inline-block; font-size: 10px; padding: 2px 7px; border-radius: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.4px; }
    .gm-iw-status.operating { background: #d1fae5; color: #047857; }
    .gm-iw-status.pipeline  { background: #fef3c7; color: #b45309; }
    .gm-iw-status.inactive  { background: #f3f4f6; color: #6b7280; }
    .gm-iw-meta { font-size: 11px; color: #556; margin: 4px 0 3px; }
    .gm-iw-stats { display: grid; grid-template-columns: auto 1fr; gap: 3px 10px; font-size: 11px; margin: 8px 0 4px; padding-top: 8px; border-top: 1px solid #eee; }
    .gm-iw-stats dt { color: #7080a0; font-weight: 600; }
    .gm-iw-stats dd { color: #1a2744; }
    .gm-iw-notes { font-size: 10px; color: #556; margin-top: 6px; line-height: 1.4; font-style: italic; }
  </style>
</head>
<body>

<div class="header">
  <div class="header-left">
    <h1>Marriott vs. Hilton &mdash; Washington, DC Hotel Saturation</h1>
    <p>__SUBTITLE__</p>
  </div>
  <div class="header-right"><a href="../">&larr; All Maps</a></div>
</div>

<div class="main">
  <div class="sidebar">
    <div class="sb-section">
      <h3>Brand family <span style="font-size:9px;color:#8090b0;text-transform:none;letter-spacing:0;font-style:italic;">&middot; click to toggle</span></h3>
      <div id="legend"></div>
    </div>
    <div class="sb-section">
      <h3>Status (CoStar) &middot; click to toggle</h3>
      <div class="filter-pills" id="status-filters"></div>
      <button class="fit-btn" id="fit-btn">Fit map to visible</button>
      <div class="filter-note" id="filter-note"></div>
    </div>
    <div class="sb-section">
      <h3>Saturation matrix &mdash; hotels by family &times; status</h3>
      <div id="matrix"></div>
    </div>
    <div class="sb-section">
      <h3>Operating room-supply scale (RBA proxy)</h3>
      <div id="rba"></div>
      <div class="matrix-note">CoStar exports building area (RBA), not room counts. Total rentable building SF across <em>operating</em> hotels is shown as a rough scale proxy for each flag's footprint.</div>
    </div>
    <div class="sb-section" style="border-bottom:none;">
      <h3>Properties &mdash; nearest to AKA first</h3>
      <div id="hotel-list"></div>
    </div>
    <div class="sidebar-footer">__FOOTER__</div>
  </div>
  <div id="map"></div>
</div>

<script>
  const HOTELS = __HOTELS_JSON__;
  const AKA = __AKA_JSON__;
  const FAM = __FAMILIES_JSON__;
  const OTHER = __OTHER_JSON__;
  const SUMMARY = __SUMMARY_JSON__;

  const STATUS_META = {
    operating: { label: 'Operating', opacity: 1.0 },
    pipeline:  { label: 'Pipeline',  opacity: 0.6 },
    inactive:  { label: 'Inactive',  opacity: 0.28 },
  };

  function famStyle(family) { return FAM[family] || OTHER; }

  let map, infoWindow;
  const markers = [];
  const listItems = [];
  const filterState = { families: new Set(['Marriott', 'Hilton', 'Other']), statuses: new Set(['operating', 'pipeline', 'inactive']) };

  function diamondIcon(family, label, status) {
    const s = famStyle(family);
    const op = STATUS_META[status].opacity;
    return {
      url: 'data:image/svg+xml;charset=UTF-8,' + encodeURIComponent(`
        <svg xmlns="http://www.w3.org/2000/svg" width="34" height="34" viewBox="0 0 34 34">
          <filter id="d"><feDropShadow dx="0" dy="2" stdDeviation="1.5" flood-opacity="0.35"/></filter>
          <rect x="6" y="6" width="22" height="22" transform="rotate(45 17 17)"
                fill="${s.color}" fill-opacity="${op}" stroke="${s.stroke}" stroke-width="2" filter="url(#d)"/>
          <text x="17" y="20" text-anchor="middle" font-family="Arial,sans-serif"
                font-size="11" font-weight="bold" fill="${s.label}">${label}</text>
        </svg>`),
      scaledSize: new google.maps.Size(34, 34),
      anchor: new google.maps.Point(17, 17),
    };
  }

  function starIcon() {
    return {
      url: 'data:image/svg+xml;charset=UTF-8,' + encodeURIComponent(`
        <svg xmlns="http://www.w3.org/2000/svg" width="46" height="46" viewBox="0 0 46 46">
          <filter id="ss"><feDropShadow dx="0" dy="2" stdDeviation="2" flood-opacity="0.45"/></filter>
          <circle cx="23" cy="23" r="15" fill="#1a2744" filter="url(#ss)"/>
          <path d="M23 11 L26.7 19 L35 20 L29 26 L30.6 34.5 L23 30.3 L15.4 34.5 L17 26 L11 20 L19.3 19 Z"
                fill="#f4c20d" stroke="#fff" stroke-width="1"/>
        </svg>`),
      scaledSize: new google.maps.Size(46, 46),
      anchor: new google.maps.Point(23, 23),
    };
  }

  function fmt(n) { return n ? Number(n).toLocaleString() : ''; }

  function buildInfoWindow(h) {
    const s = famStyle(h.family);
    const rows = [];
    if (h.star)        rows.push(`<dt>Class</dt><dd>${h.star}-Star (CoStar)</dd>`);
    if (h.rba)         rows.push(`<dt>RBA</dt><dd>${fmt(h.rba)} SF</dd>`);
    if (h.year)        rows.push(`<dt>${h.status === 'operating' ? 'Built' : 'Year'}</dt><dd>${h.year}</dd>`);
    if (h.stories)     rows.push(`<dt>Stories</dt><dd>${h.stories}</dd>`);
    rows.push(`<dt>From AKA</dt><dd>${h.dist} mi</dd>`);
    const statsBlock = `<dl class="gm-iw-stats">${rows.join('')}</dl>`;
    return `
      <div class="gm-iw">
        <div class="gm-iw-title">${h.name}</div>
        <span class="gm-iw-brand" style="background:${s.color};color:${s.label};">${h.family}</span>
        <span class="gm-iw-status ${h.status}">${h.raw_status}</span>
        ${statsBlock}
      </div>`;
  }

  function buildAkaInfoWindow() {
    return `
      <div class="gm-iw">
        <div class="gm-iw-title">&#9733; ${AKA.name}</div>
        <span class="gm-iw-brand" style="background:#f4c20d;color:#1a2744;">Subject</span>
        <div class="gm-iw-meta">${AKA.address}</div>
        <div class="gm-iw-notes">${AKA.note}</div>
      </div>`;
  }

  function buildLegend() {
    const counts = { Marriott: 0, Hilton: 0, Other: 0 };
    HOTELS.forEach(h => { counts[h.family] = (counts[h.family] || 0) + 1; });
    const order = ['Marriott', 'Hilton', 'Other'];
    const el = document.getElementById('legend');
    let html = order.filter(f => counts[f]).map(f => {
      const s = famStyle(f);
      return `<div class="legend-row clickable" data-family="${f}">
                <div class="leg-dot" style="background:${s.color};border-color:${s.stroke};"></div>
                ${f}<span class="leg-count">${counts[f]}</span>
              </div>`;
    }).join('');
    // Subject star row (non-toggle).
    html += `<div class="legend-row" style="margin-top:6px;border-top:1px solid #eef0f3;padding-top:8px;">
               <svg class="leg-star" viewBox="0 0 24 24"><path fill="currentColor" stroke="#1a2744" stroke-width="0.5" d="M12 2l3 7h7l-5.5 4.5L18.5 22 12 17.5 5.5 22l2-8.5L2 9h7z"/></svg>
               ${AKA.name} <span class="leg-count" style="color:#b8860b;">Subject</span>
             </div>`;
    el.innerHTML = html;
    el.querySelectorAll('.legend-row.clickable').forEach(row => {
      row.addEventListener('click', () => {
        const f = row.getAttribute('data-family');
        if (filterState.families.has(f)) filterState.families.delete(f);
        else filterState.families.add(f);
        row.classList.toggle('off', !filterState.families.has(f));
        applyFilters();
      });
    });
  }

  function buildStatusFilters() {
    const counts = { operating: 0, pipeline: 0, inactive: 0 };
    HOTELS.forEach(h => { counts[h.status] += 1; });
    const wrap = document.getElementById('status-filters');
    wrap.innerHTML = ['operating', 'pipeline', 'inactive'].map(s =>
      `<span class="filter-pill s-${s}" data-status="${s}">${STATUS_META[s].label} <span class="pill-count">${counts[s]}</span></span>`
    ).join('');
    wrap.querySelectorAll('.filter-pill').forEach(el => {
      el.addEventListener('click', () => {
        const s = el.getAttribute('data-status');
        if (filterState.statuses.has(s)) filterState.statuses.delete(s);
        else filterState.statuses.add(s);
        el.classList.toggle('off', !filterState.statuses.has(s));
        applyFilters();
      });
    });
    document.getElementById('fit-btn').addEventListener('click', fitVisible);
  }

  function buildMatrix() {
    const m = SUMMARY.matrix, fams = SUMMARY.fams, states = SUMMARY.states;
    const SLAB = { operating: 'Operating', pipeline: 'Pipeline', inactive: 'Inactive' };
    const cell = v => v ? v : '<span class="dash">&mdash;</span>';
    const colTot = {}; states.forEach(s => colTot[s] = 0);
    let grand = 0;
    const body = fams.filter(f => states.reduce((a, s) => a + m[f][s], 0) > 0).map(f => {
      const st = famStyle(f);
      let rowTot = 0;
      const tds = states.map(s => { rowTot += m[f][s]; colTot[s] += m[f][s]; return `<td>${cell(m[f][s])}</td>`; }).join('');
      grand += rowTot;
      return `<tr><td><span class="fam-dot" style="background:${st.color};"></span>${f}</td>${tds}<td class="tot">${rowTot}</td></tr>`;
    }).join('');
    const foot = `<tr><td>Total</td>${states.map(s => `<td>${colTot[s]}</td>`).join('')}<td>${grand}</td></tr>`;
    document.getElementById('matrix').innerHTML = `
      <table class="matrix">
        <thead><tr>
          <th class="c-fam">Family</th>
          ${states.map(s => `<th class="s-${s}">${SLAB[s]}</th>`).join('')}
          <th class="s-total">Total</th>
        </tr></thead>
        <tbody>${body}</tbody>
        <tfoot>${foot}</tfoot>
      </table>
      <div class="matrix-note">Operating = CoStar &ldquo;Existing.&rdquo; Pipeline = Under Construction + Final Planning. Inactive = Demolished / Abandoned / Deferred (incl. the demolished Marriott Wardman Park).</div>`;
  }

  function buildRba() {
    const r = SUMMARY.op_rba;
    const order = ['Marriott', 'Hilton', 'Other'].filter(f => r[f]);
    const max = Math.max(1, ...order.map(f => r[f]));
    document.getElementById('rba').innerHTML = order.map(f => {
      const st = famStyle(f);
      const mm = (r[f] / 1e6).toFixed(2);
      const pct = Math.round(100 * r[f] / max);
      return `<div class="rba-row">
                <span class="rba-lab"><span class="fam-dot" style="display:inline-block;width:9px;height:9px;transform:rotate(45deg);background:${st.color};border:1px solid rgba(0,0,0,0.15);"></span>${f}</span>
                <span style="font-variant-numeric:tabular-nums;font-weight:600;color:#1a2744;">${mm}M SF</span>
              </div>
              <div style="height:5px;background:#eef0f3;border-radius:3px;margin:1px 0 5px;overflow:hidden;"><div style="height:100%;width:${pct}%;background:${st.color};opacity:0.8;"></div></div>`;
    }).join('');
  }

  function initMap() {
    map = new google.maps.Map(document.getElementById('map'), {
      center: { lat: AKA.lat, lng: AKA.lng },
      zoom: 13,
      mapTypeControl: true,
      streetViewControl: false,
      fullscreenControl: true,
    });
    infoWindow = new google.maps.InfoWindow({ maxWidth: 310 });

    const listEl = document.getElementById('hotel-list');

    HOTELS.forEach((h, i) => {
      const marker = new google.maps.Marker({
        position: { lat: h.lat, lng: h.lng },
        map,
        icon: diamondIcon(h.family, String(h.n), h.status),
        title: h.name,
        zIndex: h.status === 'operating' ? 50 : (h.status === 'pipeline' ? 30 : 10),
      });
      markers.push(marker);
      const iw = buildInfoWindow(h);
      marker.addListener('click', () => { infoWindow.setContent(iw); infoWindow.open(map, marker); });

      const s = famStyle(h.family);
      const item = document.createElement('div');
      item.className = 'hotel-item';
      item.innerHTML = `
        <div class="h-pin" style="background:${s.color};border:1.5px solid ${s.stroke};opacity:${STATUS_META[h.status].opacity};"><span class="h-pin-num" style="color:${s.label}">${h.n}</span></div>
        <div class="h-info">
          <h4>${h.name}<span class="h-status-badge ${h.status}">${STATUS_META[h.status].label}</span></h4>
          <div class="h-meta">${h.family} &middot; ${h.dist} mi from AKA${h.rba ? ' &middot; ' + fmt(h.rba) + ' SF' : ''}</div>
        </div>`;
      item.addEventListener('click', () => {
        map.panTo({ lat: h.lat, lng: h.lng });
        map.setZoom(15);
        infoWindow.setContent(iw);
        infoWindow.open(map, marker);
      });
      listEl.appendChild(item);
      listItems.push(item);
    });

    // Starred subject — AKA White House, always on top.
    const akaMarker = new google.maps.Marker({
      position: { lat: AKA.lat, lng: AKA.lng },
      map,
      icon: starIcon(),
      title: AKA.name,
      zIndex: 999,
    });
    const akaIw = buildAkaInfoWindow();
    akaMarker.addListener('click', () => { infoWindow.setContent(akaIw); infoWindow.open(map, akaMarker); });

    buildLegend();
    buildStatusFilters();
    buildMatrix();
    buildRba();
    applyFilters();
  }

  function applyFilters() {
    let visible = 0;
    HOTELS.forEach((h, i) => {
      const on = filterState.families.has(h.family) && filterState.statuses.has(h.status);
      markers[i].setVisible(on);
      listItems[i].classList.toggle('off', !on);
      if (on) visible += 1;
    });
    const note = document.getElementById('filter-note');
    if (note) note.textContent = `${visible} of ${HOTELS.length} hotels visible (AKA subject always shown)`;
  }

  function fitVisible() {
    const b = new google.maps.LatLngBounds();
    b.extend({ lat: AKA.lat, lng: AKA.lng });
    let any = false;
    HOTELS.forEach((h, i) => { if (markers[i].getVisible()) { b.extend({ lat: h.lat, lng: h.lng }); any = true; } });
    if (any) map.fitBounds(b, { top: 60, right: 60, bottom: 60, left: 60 });
  }
</script>
<script src="https://maps.googleapis.com/maps/api/js?key=__API_KEY__&callback=initMap" async defer></script>

</body>
</html>"""


def render_html(hotels, summary):
    counts = {"Marriott": 0, "Hilton": 0, "Other": 0}
    op = {"Marriott": 0, "Hilton": 0, "Other": 0}
    for h in hotels:
        counts[h["family"]] += 1
        if h["status"] == "operating":
            op[h["family"]] += 1
    subtitle = (
        f"CoStar export &middot; {len(hotels)} properties &middot; "
        f"Marriott {counts['Marriott']} ({op['Marriott']} operating) "
        f"vs. Hilton {counts['Hilton']} ({op['Hilton']} operating) "
        f"&middot; Subject: AKA White House"
    )
    footer = (
        "Universe = CoStar hotel export for the Washington DC CBD submarket (Marriott &amp; "
        "Hilton brand families plus citizenM); not every DC hotel. Brand family inferred "
        "from flag name; status and building area (RBA) per CoStar. The AKA White House "
        "(independent serviced residence) is the starred subject, not part of either flag."
    )

    fam_js = {f: {"color": c["color"], "stroke": c["stroke"], "label": c["label"]}
              for f, c in FAMILIES.items()}

    html = TEMPLATE
    html = html.replace("__SUBTITLE__", subtitle)
    html = html.replace("__FOOTER__", footer)
    html = html.replace("__HOTELS_JSON__", json.dumps(hotels, indent=2))
    html = html.replace("__AKA_JSON__", json.dumps(AKA))
    html = html.replace("__FAMILIES_JSON__", json.dumps(fam_js))
    html = html.replace("__OTHER_JSON__", json.dumps({"color": OTHER["color"], "stroke": OTHER["stroke"], "label": OTHER["label"]}))
    html = html.replace("__SUMMARY_JSON__", json.dumps(summary))
    html = html.replace("__API_KEY__", API_KEY)
    return html


def update_root_index(hotels):
    index_path = os.path.join(REPO_DIR, "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        content = f.read()

    href = f"./{SLUG}/"
    if f'href="{href}"' in content:
        print("  Already in root index.html")
        return

    counts = {"Marriott": 0, "Hilton": 0}
    for h in hotels:
        if h["family"] in counts:
            counts[h["family"]] += 1
    entry = f"""      <li>
        <a href="{href}">
          Marriott vs. Hilton &mdash; Washington, DC Saturation
          <div class="meta">Market saturation analysis &middot; {len(hotels)} hotels (Marriott {counts['Marriott']} / Hilton {counts['Hilton']}) &middot; Subject: AKA White House</div>
        </a>
      </li>"""

    if "<!-- MAPS_END -->" not in content:
        print("  WARNING: <!-- MAPS_END --> marker missing; skipping index insert")
        return
    content = content.replace("      <!-- MAPS_END -->", f"{entry}\n      <!-- MAPS_END -->")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(content)
    print("  Added to root index.html")


def main():
    parser = argparse.ArgumentParser(description="Generate the DC Marriott-vs-Hilton saturation map")
    parser.add_argument("--deploy", action="store_true", help="Push to GitHub Pages and verify")
    args = parser.parse_args()

    if not os.path.exists(SRC):
        print(f"Source not found: {SRC}")
        sys.exit(1)

    print(f"Reading {SRC}")
    hotels = load_hotels()
    summary = summarize(hotels)

    fams = {"Marriott": 0, "Hilton": 0, "Other": 0}
    for h in hotels:
        fams[h["family"]] += 1
    print(f"  {len(hotels)} hotels: Marriott {fams['Marriott']}, Hilton {fams['Hilton']}, Other {fams['Other']}")
    print("  Saturation matrix (hotels):")
    print("    family     operating  pipeline  inactive")
    for f in summary["fams"]:
        m = summary["matrix"][f]
        print(f"    {f:<10} {m['operating']:>9} {m['pipeline']:>9} {m['inactive']:>9}")

    html = render_html(hotels, summary)
    out_dir = os.path.join(REPO_DIR, SLUG)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "index.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Saved: {out_path}")

    update_root_index(hotels)

    url = f"{PAGES_BASE_URL}/{SLUG}/"
    if not args.deploy:
        print(f"\nGenerated. To deploy: python generate_dc_saturation_map.py --deploy")
        print(f"  Will be live at: {url}")
        return

    print("\nDeploying to GitHub Pages...")
    _commit_and_push("Add Marriott vs. Hilton DC saturation map (AKA White House subject)")
    ok = verify_deploy([url])
    print("\nDEPLOY VERIFIED [OK]" if ok else "DEPLOY INCOMPLETE — see warnings above")
    print(f"  {url}")
    if not ok:
        sys.exit(2)


if __name__ == "__main__":
    main()
